import openpyxl
import os


file_path = os.path.dirname(os.path.abspath(__file__))
base_path = os.path.join(file_path, 'openpyxl1_data.xlsx')


workbook = openpyxl.load_workbook(base_path)
worksheet = workbook['Sheet1']

row3=[item.value for item in list(worksheet.rows)[2]]
print('第3行值',row3)


col3=[item.value for item in list(worksheet.columns)[2]]
print('第3列值',col3)


cell_2_3=worksheet.cell(row=2,column=3).value
print('第2行第3列值',cell_2_3)


max_row=worksheet.max_row
print('最大行',max_row)


