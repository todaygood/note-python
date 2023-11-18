import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "test"

print("1 ws1.cell(10,3):",ws1.cell(10,3).value)

ws1.cell(column=10, row=3, value='555')
wb.save(filename = dest_filename)

wb=openpyxl.load_workbook(dest_filename)
ws2=wb['test']
print("2 ws2.cell(10,3):",ws2.cell(10,3).value)





'''

print("1 ws1.cell(10,3):",ws1.cell(10,3).value)
value="{0}".format(get_column_letter(10))
print(value)


for row in range(1, 40):
     ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")



print(ws3['AA10'].value)

wb.save(filename = dest_filename)

print("3 ws1.cell(10,3):",ws1.cell(10,3).value)
'''